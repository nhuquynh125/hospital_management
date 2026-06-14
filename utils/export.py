"""
Hospital Management System — Export Utility
Xuất PDF (reportlab) và Excel (openpyxl)
"""

import os
from datetime import datetime

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB = True
except ImportError:
    REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL = True
except ImportError:
    OPENPYXL = False


EXPORT_DIR = os.path.join(os.path.expanduser("~"), "Downloads", "HospitalExports")


def ensure_export_dir():
    os.makedirs(EXPORT_DIR, exist_ok=True)
    return EXPORT_DIR


# ═══════════════════════════════════════════════════════════
#  PDF Export
# ═══════════════════════════════════════════════════════════
def export_patients_pdf(patients: list) -> str:
    """Export patient list to PDF. Returns file path."""
    if not REPORTLAB:
        raise ImportError("reportlab chưa được cài. Chạy: pip install reportlab")

    ensure_export_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(EXPORT_DIR, f"DanhSachBenhNhan_{timestamp}.pdf")

    doc = SimpleDocTemplate(filepath, pagesize=A4,
                             leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                  fontSize=16, spaceAfter=4, alignment=1)
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                  fontSize=10, spaceAfter=12, alignment=1, textColor=colors.grey)
    story.append(Paragraph("🏥 BỆNH VIỆN QUẢN LÝ HỆ THỐNG", title_style))
    story.append(Paragraph(f"DANH SÁCH BỆNH NHÂN — Xuất ngày {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2b6cb0")))
    story.append(Spacer(1, 0.4*cm))

    # Table
    headers = ["Mã BN", "Họ tên", "Ngày sinh", "Giới tính", "Điện thoại", "Nhóm máu", "Số BHYT"]
    data = [headers]
    for p in patients:
        data.append([
            p["patient_code"] or "",
            p["full_name"] or "",
            p["birth_date"] or "",
            p["gender"] or "",
            p["phone"] or "",
            p["blood_type"] or "",
            p["insurance_id"] or "",
        ])

    col_widths = [2.2*cm, 5*cm, 2.8*cm, 2.4*cm, 3*cm, 2.4*cm, 3.2*cm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#2b6cb0")),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("FONTSIZE",    (0, 0), (-1, 0),  9),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("ALIGN",       (1, 1), (1, -1),  "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e0")),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0,0), (-1, -1), 5),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(f"Tổng cộng: {len(patients)} bệnh nhân", styles["Normal"]))

    doc.build(story)
    return filepath


def export_prescription_pdf(prescription: dict, items: list, patient_name: str, doctor_name: str) -> str:
    """Export a prescription to PDF."""
    if not REPORTLAB:
        raise ImportError("reportlab chưa được cài. Chạy: pip install reportlab")

    ensure_export_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(EXPORT_DIR, f"DonThuoc_{timestamp}.pdf")

    doc = SimpleDocTemplate(filepath, pagesize=A4,
                             leftMargin=2.5*cm, rightMargin=2.5*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    title_style = ParagraphStyle("t", parent=styles["Heading1"], fontSize=18, alignment=1, spaceAfter=2)
    sub_style   = ParagraphStyle("s", parent=styles["Normal"], fontSize=10, alignment=1,
                                  textColor=colors.grey, spaceAfter=10)
    story.append(Paragraph("🏥 ĐƠN THUỐC", title_style))
    story.append(Paragraph("HOSPITAL MANAGEMENT SYSTEM", sub_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#553c9a")))
    story.append(Spacer(1, 0.3*cm))

    # Patient & doctor info
    info_data = [
        ["Bệnh nhân:", patient_name,  "Bác sĩ:", doctor_name],
        ["Ngày kê đơn:", datetime.now().strftime("%d/%m/%Y"), "Mã đơn:", f"#{prescription.get('id', 'N/A')}"],
    ]
    info_tbl = Table(info_data, colWidths=[3*cm, 7*cm, 2.5*cm, 4*cm])
    info_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0,0),(-1,-1), 10),
        ("FONTNAME", (0,0),(0,-1), "Helvetica-Bold"),
        ("FONTNAME", (2,0),(2,-1), "Helvetica-Bold"),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 0.4*cm))

    # Medicines table
    story.append(Paragraph("CHI TIẾT ĐƠN THUỐC:", ParagraphStyle("h", parent=styles["Heading3"], fontSize=11)))
    story.append(Spacer(1, 0.2*cm))
    med_headers = ["STT", "Tên thuốc", "Số lượng", "Liều dùng", "Số ngày", "Ghi chú"]
    med_data = [med_headers]
    for i, item in enumerate(items, 1):
        med_data.append([
            str(i), item.get("name",""), str(item.get("quantity","")),
            item.get("dosage",""), f"{item.get('duration_days','')} ngày",
            item.get("notes","")
        ])
    med_tbl = Table(med_data, colWidths=[1*cm, 5*cm, 2*cm, 4.5*cm, 2*cm, 3*cm], repeatRows=1)
    med_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0),(-1,0),  colors.HexColor("#553c9a")),
        ("TEXTCOLOR",   (0,0),(-1,0),  colors.white),
        ("FONTSIZE",    (0,0),(-1,-1), 9),
        ("ALIGN",       (0,0),(-1,-1), "CENTER"),
        ("ALIGN",       (1,1),(1,-1),  "LEFT"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#f5f3ff")]),
        ("GRID",        (0,0),(-1,-1), 0.5, colors.HexColor("#e9d8fd")),
        ("TOPPADDING",  (0,0),(-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))
    story.append(med_tbl)

    if prescription.get("notes"):
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(f"<b>Lưu ý:</b> {prescription['notes']}", styles["Normal"]))

    story.append(Spacer(1, 1*cm))
    sig_data = [["", "Chữ ký bác sĩ"], ["", doctor_name]]
    sig_tbl = Table(sig_data, colWidths=[12*cm, 5*cm])
    sig_tbl.setStyle(TableStyle([
        ("ALIGN", (1,0),(1,-1), "CENTER"),
        ("FONTSIZE", (0,0),(-1,-1), 10),
    ]))
    story.append(sig_tbl)

    doc.build(story)
    return filepath


# ═══════════════════════════════════════════════════════════
#  Excel Export
# ═══════════════════════════════════════════════════════════
def export_patients_excel(patients: list) -> str:
    if not OPENPYXL:
        raise ImportError("openpyxl chưa được cài. Chạy: pip install openpyxl")

    ensure_export_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(EXPORT_DIR, f"DanhSachBenhNhan_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách bệnh nhân"

    # Header style
    header_fill   = PatternFill("solid", fgColor="2B6CB0")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    alt_fill = PatternFill("solid", fgColor="EBF8FF")

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = f"DANH SÁCH BỆNH NHÂN — Xuất ngày {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font      = Font(bold=True, size=14, color="1A365D")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = ["Mã BN", "Họ tên", "Ngày sinh", "Giới tính",
               "Điện thoại", "Nhóm máu", "Số BHYT", "Ngày tạo"]
    col_widths = [12, 25, 13, 12, 16, 11, 18, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    # Data rows
    for r_idx, p in enumerate(patients, 3):
        row_vals = [
            p["patient_code"], p["full_name"], p["birth_date"] or "",
            p["gender"] or "", p["phone"] or "", p["blood_type"] or "",
            p["insurance_id"] or "", (p["created_at"] or "")[:10]
        ]
        fill = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r_idx].height = 16

    # Summary row
    last_row = len(patients) + 3
    ws.cell(row=last_row, column=1, value=f"Tổng cộng: {len(patients)} bệnh nhân").font = Font(bold=True, italic=True)
    ws.freeze_panes = "A3"

    wb.save(filepath)
    return filepath


def export_monthly_report_excel(stats: dict, appointments: list) -> str:
    if not OPENPYXL:
        raise ImportError("openpyxl chưa được cài.")

    ensure_export_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(EXPORT_DIR, f"BaoCaoThang_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo tháng"

    ws["A1"] = "BÁO CÁO TỔNG HỢP THÁNG"
    ws["A1"].font = Font(bold=True, size=16, color="1A365D")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:D1")

    ws["A3"] = "Tổng bệnh nhân:"
    ws["B3"] = stats.get("total_patients", 0)
    ws["A4"] = "Tổng nhân viên:"
    ws["B4"] = stats.get("total_staff", 0)
    ws["A5"] = "Lịch hẹn hôm nay:"
    ws["B5"] = stats.get("today_appointments", 0)
    ws["A6"] = "Phòng trống:"
    ws["B6"] = stats.get("available_rooms", 0)

    for row in ws.iter_rows(min_row=3, max_row=6, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

    wb.save(filepath)
    return filepath
